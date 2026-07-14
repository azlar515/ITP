import json
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from .models import AuditLog, ItpItem, ItpVersion, ItpVersionItem, Project


@dataclass
class ParsedRow:
    sort_order: int
    parent_code: str | None
    code: str
    title_zh: str | None
    title_en: str
    item_uid: str | None = None
    before_sea_trial: bool = False


def normalize_cell(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan", "null"}:
        return None
    return text


def is_header_row(row: tuple[Any, ...]) -> bool:
    normalized = [normalize_cell(cell) for cell in row[:5]]
    lowered = {cell.lower() for cell in normalized if cell}
    parent_label = (normalized[1] or "").lower() if len(normalized) > 1 else ""
    code_label = (normalized[2] or "").lower() if len(normalized) > 2 else ""
    if parent_label in {"上级编码", "parent code", "parent"} and code_label in {"当前编码", "current code", "code"}:
        return True
    header_words = {
        "no",
        "no.",
        "序号",
        "parent code",
        "parent",
        "上级编码",
        "current code",
        "code",
        "当前编码",
        "中文描述",
        "chinese description",
        "英文描述",
        "english description",
    }
    return len(lowered.intersection(header_words)) >= 2


def looks_like_code(value: Any, *, allow_project_name: bool = False) -> bool:
    text = normalize_cell(value)
    if not text:
        return False
    if "\n" in text or len(text) > 80:
        return False
    if not any(char.isalpha() for char in text):
        return False
    if allow_project_name:
        return True
    return " " not in text


def worksheet_score(sheet: Worksheet) -> int:
    score = 0
    max_scan_rows = min(sheet.max_row, 250)
    for row in sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True):
        if len(row) < 3 or is_header_row(row):
            continue
        parent_code = row[1] if len(row) > 1 else None
        code = row[2] if len(row) > 2 else None
        if looks_like_code(parent_code, allow_project_name=True) and looks_like_code(code):
            score += 1
    if sheet.sheet_state == "visible":
        score += 25
    return score


def select_itp_sheet(workbook) -> Worksheet:
    candidates = sorted(workbook.worksheets, key=worksheet_score, reverse=True)
    if not candidates or worksheet_score(candidates[0]) <= 0:
        raise ValueError("No worksheet looks like an ITP import sheet. Expected parent code in column B and current code in column C.")
    return candidates[0]


def parse_excel_itp(content: bytes) -> tuple[str, list[ParsedRow]]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = select_itp_sheet(workbook)
    rows: list[ParsedRow] = []

    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not any(normalize_cell(cell) for cell in row):
            continue
        if is_header_row(row):
            continue
        parent_code = normalize_cell(row[1] if len(row) > 1 else None)
        code = normalize_cell(row[2] if len(row) > 2 else None)
        if not code:
            continue
        title_zh = normalize_cell(row[3] if len(row) > 3 else None)
        title_en = normalize_cell(row[4] if len(row) > 4 else None) or code
        item_uid = normalize_cell(row[5] if len(row) > 5 else None)
        marker = (normalize_cell(row[6] if len(row) > 6 else None) or "").lower()
        before_sea_trial = marker in {"y", "yes", "true", "1", "items before sea trial"}
        rows.append(ParsedRow(sort_order=idx, parent_code=parent_code, code=code, title_zh=title_zh, title_en=title_en, item_uid=item_uid, before_sea_trial=before_sea_trial))

    if not rows:
        raise ValueError("No valid ITP rows found. Expected code in the third column.")

    root_candidates = [row for row in rows if row.parent_code is None]
    project_name = root_candidates[0].code if root_candidates else rows[0].parent_code or "Unnamed Project"
    if project_name and all(row.code != project_name for row in rows):
        rows.insert(
            0,
            ParsedRow(
                sort_order=0,
                parent_code=None,
                code=project_name,
                title_zh=None,
                title_en=project_name,
                item_uid=None,
                before_sea_trial=False,
            ),
        )
    return project_name, rows


def snapshot_item(item: ItpItem) -> dict[str, Any]:
    return {
        "id": item.id,
        "project_id": item.project_id,
        "item_uid": item.item_uid,
        "parent_uid": item.parent_uid,
        "version_id": item.version_id,
        "parent_code": item.parent_code,
        "code": item.code,
        "title_zh": item.title_zh,
        "title_en": item.title_en,
        "level": item.level,
        "is_inspection": item.is_inspection,
        "before_sea_trial": item.before_sea_trial,
        "active": item.active,
        "sort_order": item.sort_order,
    }


def write_audit(
    db: Session,
    *,
    entity_type: str,
    entity_id: int | None,
    action: str,
    summary: str,
    actor: str,
    before: dict[str, Any] | None = None,
    after: dict[str, Any] | None = None,
) -> None:
    db.add(
        AuditLog(
            entity_type=entity_type,
            entity_id=entity_id,
            action=action,
            summary=summary,
            actor=actor,
            before_json=json.dumps(before, ensure_ascii=True, default=str) if before else None,
            after_json=json.dumps(after, ensure_ascii=True, default=str) if after else None,
        )
    )


def find_or_create_project(db: Session, name: str, actor: str) -> Project:
    project = db.query(Project).filter(Project.name == name).one_or_none()
    if project:
        return project
    project = Project(name=name)
    db.add(project)
    db.flush()
    write_audit(
        db,
        entity_type="project",
        entity_id=project.id,
        action="create",
        summary=f"Created project {name}",
        actor=actor,
        after={"id": project.id, "name": project.name},
    )
    return project


def next_version_no(db: Session, project_id: int) -> int:
    latest = db.query(ItpVersion).filter(ItpVersion.project_id == project_id).order_by(ItpVersion.version_no.desc()).first()
    return (latest.version_no + 1) if latest else 1


def resolve_levels_and_leaf_flags(items: list[ItpItem]) -> None:
    by_code = {item.code: item for item in items}
    child_parent_codes = {item.parent_code for item in items if item.parent_code}

    def level_for(item: ItpItem, seen: set[str] | None = None) -> int:
        seen = seen or set()
        if not item.parent_code or item.parent_code not in by_code or item.parent_code in seen:
            return 1
        seen.add(item.code)
        return level_for(by_code[item.parent_code], seen) + 1

    for item in items:
        item.parent_id = by_code[item.parent_code].id if item.parent_code in by_code else None
        item.parent_uid = by_code[item.parent_code].item_uid if item.parent_code in by_code else None
        item.level = level_for(item)
        item.is_inspection = item.level >= 5 and item.code not in child_parent_codes


def create_itp_version_snapshot(db: Session, project: Project, actor: str, source: str) -> ItpVersion:
    version = ItpVersion(project_id=project.id, version_no=next_version_no(db, project.id), source=source, created_by=actor)
    db.add(version)
    db.flush()
    items = db.query(ItpItem).filter(ItpItem.project_id == project.id).order_by(ItpItem.sort_order, ItpItem.code).all()
    for item in items:
        item.version_id = version.id
        db.add(
            ItpVersionItem(
                version_id=version.id,
                project_id=project.id,
                item_uid=item.item_uid,
                parent_uid=item.parent_uid,
                parent_code=item.parent_code,
                code=item.code,
                title_zh=item.title_zh,
                title_en=item.title_en,
                level=item.level,
                is_inspection=item.is_inspection,
                before_sea_trial=item.before_sea_trial,
                active=item.active,
                sort_order=item.sort_order,
            )
        )
    project.active_itp_version_id = version.id
    write_audit(
        db,
        entity_type="itp_version",
        entity_id=version.id,
        action="create",
        summary=f"Created ITP version v{version.version_no} for project {project.name}",
        actor=actor,
        after={"project_id": project.id, "version_id": version.id, "version_no": version.version_no, "item_count": len(items)},
    )
    return version


def preview_import(db: Session, content: bytes) -> dict[str, Any]:
    project_name, rows = parse_excel_itp(content)
    project = db.query(Project).filter(Project.name == project_name).one_or_none()
    existing = {}
    if project:
        existing_by_code = {item.code: item for item in db.query(ItpItem).filter(ItpItem.project_id == project.id).all()}
        existing_by_uid = {item.item_uid: item for item in existing_by_code.values()}
        existing = {"by_code": existing_by_code, "by_uid": existing_by_uid}

    creates: list[dict[str, Any]] = []
    updates: list[dict[str, Any]] = []
    unchanged = 0
    uploaded_codes = {row.code for row in rows}

    for row in rows:
        item = None
        if existing:
            item = existing["by_uid"].get(row.item_uid) if row.item_uid else None
            item = item or existing["by_code"].get(row.code)
        incoming = {
            "code": row.code,
            "parent_code": row.parent_code,
            "title_zh": row.title_zh,
            "title_en": row.title_en,
            "sort_order": row.sort_order,
            "before_sea_trial": row.before_sea_trial,
            "active": True,
        }
        if not item:
            creates.append(incoming)
            continue
        changed = {}
        for key, value in incoming.items():
            if getattr(item, key) != value:
                changed[key] = {"from": getattr(item, key), "to": value}
        if changed:
            updates.append({"code": row.code, "changes": changed})
        else:
            unchanged += 1

    existing_items = existing["by_code"] if existing else {}
    missing = [
        {"code": item.code, "title_zh": item.title_zh, "title_en": item.title_en, "parent_code": item.parent_code, "item_uid": item.item_uid, "active": item.active}
        for code, item in existing_items.items()
        if code not in uploaded_codes
    ]

    return {
        "project_name": project_name,
        "rows": len(rows),
        "creates": creates,
        "updates": updates,
        "unchanged": unchanged,
        "missing_from_upload": missing,
    }


def import_itp(db: Session, content: bytes, actor: str, mode: str = "partial") -> Project:
    if mode not in {"partial", "global"}:
        raise ValueError("Import mode must be partial or global.")
    project_name, rows = parse_excel_itp(content)
    project = find_or_create_project(db, project_name, actor)
    existing_items = db.query(ItpItem).filter(ItpItem.project_id == project.id).all()
    existing_by_code = {item.code: item for item in existing_items}
    existing_by_uid = {item.item_uid: item for item in existing_items}
    original_snapshots = {item.id: snapshot_item(item) for item in existing_items}
    uploaded_item_ids: set[int] = set()
    incoming_codes = [row.code for row in rows]
    duplicate_codes = sorted({code for code in incoming_codes if incoming_codes.count(code) > 1})
    if duplicate_codes:
        raise ValueError(f"Duplicate Current Code in upload: {', '.join(duplicate_codes[:10])}")

    matched: list[tuple[ParsedRow, ItpItem | None]] = []
    target_code_owner: dict[str, ItpItem] = {}
    matched_item_ids: set[int] = set()
    for row in rows:
        item = existing_by_uid.get(row.item_uid) if row.item_uid else None
        item = item or existing_by_code.get(row.code)
        if item is not None:
            if item.id in matched_item_ids:
                raise ValueError(f"The same existing ITP item is matched more than once. Check UID/code near {row.code}.")
            matched_item_ids.add(item.id)
            owner = target_code_owner.get(row.code)
            if owner is not None and owner.id != item.id:
                raise ValueError(f"Multiple ITP items would use Current Code {row.code}.")
            target_code_owner[row.code] = item
        matched.append((row, item))

    changing_items = [item for row, item in matched if item is not None and item.code != row.code]
    incoming_code_set = {row.code for row in rows}
    blocked_missing_items = [
        item
        for item in existing_items
        if item.id not in {matched_item.id for _, matched_item in matched if matched_item is not None}
        and item.code in incoming_code_set
    ]
    if mode == "partial" and blocked_missing_items:
        blocked_codes = ", ".join(item.code for item in blocked_missing_items[:10])
        raise ValueError(f"Partial update cannot reuse code from another existing item: {blocked_codes}")
    for item in changing_items + blocked_missing_items:
        item.code = f"__tmp_import_{item.id}_{uuid4().hex[:8]}"
    if changing_items or blocked_missing_items:
        db.flush()

    for row, item in matched:
        if item is None:
            item = ItpItem(
                project_id=project.id,
                item_uid=row.item_uid or str(uuid4()),
                parent_code=row.parent_code,
                code=row.code,
                title_zh=row.title_zh,
                title_en=row.title_en,
                sort_order=row.sort_order,
                before_sea_trial=row.before_sea_trial,
                active=True,
            )
            db.add(item)
            db.flush()
            uploaded_item_ids.add(item.id)
            write_audit(
                db,
                entity_type="itp_item",
                entity_id=item.id,
                action="create",
                summary=f"Created ITP item {row.code}",
                actor=actor,
                after=snapshot_item(item),
            )
            continue

        before = original_snapshots.get(item.id, snapshot_item(item))
        item.parent_code = row.parent_code
        item.title_zh = row.title_zh
        item.title_en = row.title_en
        item.code = row.code
        item.sort_order = row.sort_order
        item.before_sea_trial = row.before_sea_trial
        item.active = True
        item.updated_at = datetime.utcnow()
        uploaded_item_ids.add(item.id)
        after = snapshot_item(item)
        if before != after:
            write_audit(
                db,
                entity_type="itp_item",
                entity_id=item.id,
                action="update",
                summary=f"Updated ITP item {row.code}",
                actor=actor,
                before=before,
                after=after,
            )

    if mode == "global":
        for item in existing_items:
            if item.id in uploaded_item_ids or not item.active:
                continue
            before = original_snapshots.get(item.id, snapshot_item(item))
            if item in blocked_missing_items:
                item.code = f"__inactive_{item.id}_{before['code']}"[:120]
            item.active = False
            item.updated_at = datetime.utcnow()
            write_audit(
                db,
                entity_type="itp_item",
                entity_id=item.id,
                action="deactivate",
                summary=f"Deactivated missing ITP item {item.code} during global import",
                actor=actor,
                before=before,
                after=snapshot_item(item),
            )

    db.flush()
    items = db.query(ItpItem).filter(ItpItem.project_id == project.id).all()
    before_leaf = {item.id: snapshot_item(item) for item in items}
    resolve_levels_and_leaf_flags(items)
    for item in items:
        before = before_leaf[item.id]
        after = snapshot_item(item)
        if before != after:
            write_audit(
                db,
                entity_type="itp_item",
                entity_id=item.id,
                action="classify",
                summary=f"Resolved hierarchy for {item.code}",
                actor=actor,
                before=before,
                after=after,
            )
    create_itp_version_snapshot(db, project, actor, f"excel_import_{mode}")
    db.commit()
    db.refresh(project)
    return project
