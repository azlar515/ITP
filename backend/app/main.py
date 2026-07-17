import base64
import hashlib
import hmac
import json
import os
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from uuid import uuid4

from fastapi import Depends, FastAPI, File, Header, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image as ReportLabImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from .database import Base, SessionLocal, engine, get_db
from .models import AuditLog, ItpItem, ItpVersionItem, Project, Ship, ShipProgress, ShipProgressEvent
from .schemas import (
    AuditLogOut,
    ImportPreview,
    ItpItemCreate,
    ItpItemOut,
    ItpItemUpdate,
    LoginRequest,
    LoginResponse,
    ProgressUpdate,
    ProjectCreate,
    ProjectOut,
    ShipCreate,
    ShipOut,
    ShipProgressOut,
    ShipUpdate,
)
from .services import create_itp_version_snapshot, import_itp, preview_import, resolve_levels_and_leaf_flags, snapshot_item, write_audit

Base.metadata.create_all(bind=engine)

with engine.begin() as connection:
    project_columns = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(projects)").fetchall()}
    if "active_itp_version_id" not in project_columns:
        connection.exec_driver_sql("ALTER TABLE projects ADD COLUMN active_itp_version_id INTEGER")

    columns = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(itp_items)").fetchall()}
    if "title_zh" not in columns:
        connection.exec_driver_sql("ALTER TABLE itp_items ADD COLUMN title_zh VARCHAR(260)")
    if "title_en" not in columns:
        connection.exec_driver_sql("ALTER TABLE itp_items ADD COLUMN title_en VARCHAR(260) DEFAULT '' NOT NULL")
    if "item_uid" not in columns:
        connection.exec_driver_sql("ALTER TABLE itp_items ADD COLUMN item_uid VARCHAR(36)")
    if "parent_uid" not in columns:
        connection.exec_driver_sql("ALTER TABLE itp_items ADD COLUMN parent_uid VARCHAR(36)")
    if "version_id" not in columns:
        connection.exec_driver_sql("ALTER TABLE itp_items ADD COLUMN version_id INTEGER")
    if "before_sea_trial" not in columns:
        connection.exec_driver_sql("ALTER TABLE itp_items ADD COLUMN before_sea_trial BOOLEAN DEFAULT 0 NOT NULL")
    if "active" not in columns:
        connection.exec_driver_sql("ALTER TABLE itp_items ADD COLUMN active BOOLEAN DEFAULT 1 NOT NULL")

    version_item_columns = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(itp_version_items)").fetchall()}
    if "before_sea_trial" not in version_item_columns:
        connection.exec_driver_sql("ALTER TABLE itp_version_items ADD COLUMN before_sea_trial BOOLEAN DEFAULT 0 NOT NULL")
    if "active" not in version_item_columns:
        connection.exec_driver_sql("ALTER TABLE itp_version_items ADD COLUMN active BOOLEAN DEFAULT 1 NOT NULL")

    progress_columns = {row[1] for row in connection.exec_driver_sql("PRAGMA table_info(ship_progress)").fetchall()}
    progress_migrations = {
        "item_uid": "ALTER TABLE ship_progress ADD COLUMN item_uid VARCHAR(36)",
        "version_id": "ALTER TABLE ship_progress ADD COLUMN version_id INTEGER",
        "code_snapshot": "ALTER TABLE ship_progress ADD COLUMN code_snapshot VARCHAR(120)",
        "parent_code_snapshot": "ALTER TABLE ship_progress ADD COLUMN parent_code_snapshot VARCHAR(120)",
        "title_zh_snapshot": "ALTER TABLE ship_progress ADD COLUMN title_zh_snapshot VARCHAR(260)",
        "title_en_snapshot": "ALTER TABLE ship_progress ADD COLUMN title_en_snapshot VARCHAR(260)",
        "revision": "ALTER TABLE ship_progress ADD COLUMN revision INTEGER DEFAULT 0 NOT NULL",
        "last_event_id": "ALTER TABLE ship_progress ADD COLUMN last_event_id INTEGER",
    }
    for column_name, ddl in progress_migrations.items():
        if column_name not in progress_columns:
            connection.exec_driver_sql(ddl)

with SessionLocal() as migration_db:
    for item in migration_db.query(ItpItem).filter((ItpItem.item_uid == None) | (ItpItem.item_uid == "")).all():
        item.item_uid = str(uuid4())
    migration_db.flush()
    items = migration_db.query(ItpItem).all()
    by_project: dict[int, list[ItpItem]] = {}
    for item in items:
        by_project.setdefault(item.project_id, []).append(item)
    for project_items in by_project.values():
        resolve_levels_and_leaf_flags(project_items)
    for progress in migration_db.query(ShipProgress).filter(ShipProgress.item_uid == None).all():
        item = migration_db.get(ItpItem, progress.itp_item_id)
        if item:
            progress.item_uid = item.item_uid
            progress.version_id = item.version_id
            progress.code_snapshot = item.code
            progress.parent_code_snapshot = item.parent_code
            progress.title_zh_snapshot = item.title_zh
            progress.title_en_snapshot = item.title_en
    migration_db.commit()

app = FastAPI(title="Shipyard ITP Database")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


USER_PASSWORD = os.environ.get("ITP_USER_PASSWORD", "0000")
ADMIN_PASSWORD = os.environ.get("ITP_ADMIN_PASSWORD", "Admin")
AUTH_SECRET = os.environ.get("ITP_AUTH_SECRET", "change-this-secret-for-production")


def sign_token(payload: str) -> str:
    return hmac.new(AUTH_SECRET.encode(), payload.encode(), hashlib.sha256).hexdigest()


def create_token(role: str, user: str) -> str:
    payload = f"{role}:{user}"
    encoded = base64.urlsafe_b64encode(payload.encode()).decode().rstrip("=")
    return f"{encoded}.{sign_token(encoded)}"


def parse_token(token: str) -> dict[str, str] | None:
    try:
        encoded, signature = token.split(".", 1)
        if not hmac.compare_digest(signature, sign_token(encoded)):
            return None
        padded = encoded + "=" * (-len(encoded) % 4)
        role, user = base64.urlsafe_b64decode(padded.encode()).decode().split(":", 1)
        if role not in {"user", "admin"}:
            return None
        return {"role": role, "user": user}
    except Exception:
        return None


def current_session(authorization: str = Header(default="")) -> dict[str, str]:
    prefix = "Bearer "
    if not authorization.startswith(prefix):
        raise HTTPException(status_code=401, detail="Authentication required.")
    session = parse_token(authorization[len(prefix) :])
    if session is None:
        raise HTTPException(status_code=401, detail="Invalid authentication token.")
    return session


def require_admin(session: dict[str, str] = Depends(current_session)) -> None:
    if session["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin role required.")


def actor(session: dict[str, str] = Depends(current_session)) -> str:
    return f"{session['user']} ({session['role']})"


@app.post("/api/auth/login", response_model=LoginResponse)
def login(payload: LoginRequest):
    if hmac.compare_digest(payload.password, ADMIN_PASSWORD):
        return LoginResponse(role="admin", user="admin", token=create_token("admin", "admin"))
    if hmac.compare_digest(payload.password, USER_PASSWORD):
        return LoginResponse(role="user", user="user", token=create_token("user", "user"))
    raise HTTPException(status_code=401, detail="Invalid password.")


CODE_PART_RE = re.compile(r"\d+|\D+")
PDF_FONT_CACHE: str | None = None


def natural_code_key(code: str | None) -> tuple:
    parts = CODE_PART_RE.findall(code or "")
    return tuple((0, int(part)) if part.isdigit() else (1, part.lower()) for part in parts)


def item_code_key(item: ItpItem | ItpItemOut) -> tuple:
    return natural_code_key(item.code)


def sort_tree_nodes(nodes: list[ItpItemOut]) -> list[ItpItemOut]:
    nodes.sort(key=item_code_key)
    for node in nodes:
        sort_tree_nodes(node.children)
    return nodes


def build_tree(items: list[ItpItem]) -> list[ItpItemOut]:
    out_by_id = {
        item.id: ItpItemOut(
            id=item.id,
            project_id=item.project_id,
            parent_id=item.parent_id,
            parent_code=item.parent_code,
            code=item.code,
            title_zh=item.title_zh,
            title_en=item.title_en,
            level=item.level,
            is_inspection=item.is_inspection,
            before_sea_trial=item.before_sea_trial,
            active=item.active,
            sort_order=item.sort_order,
            children=[],
        )
        for item in items
    }
    roots: list[ItpItemOut] = []
    for item in sorted(items, key=lambda row: (row.level, row.sort_order, row.code)):
        node = out_by_id[item.id]
        if item.parent_id and item.parent_id in out_by_id:
            out_by_id[item.parent_id].children.append(node)
        else:
            roots.append(node)
    return sort_tree_nodes(roots)


def pdf_font_name() -> str:
    global PDF_FONT_CACHE
    if PDF_FONT_CACHE:
        return PDF_FONT_CACHE
    candidates = [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simsun.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("ITPFont", path))
                PDF_FONT_CACHE = "ITPFont"
                return "ITPFont"
            except Exception:
                continue
    PDF_FONT_CACHE = "Helvetica"
    return "Helvetica"


def pg_logo_path() -> str | None:
    path = Path(__file__).resolve().parents[2] / "frontend" / "public" / "pg-logo.png"
    return str(path) if path.exists() else None


def apply_item_snapshot(item: ItpItem, snapshot: dict) -> None:
    item.project_id = snapshot["project_id"]
    item.item_uid = snapshot.get("item_uid", item.item_uid)
    item.parent_uid = snapshot.get("parent_uid")
    item.version_id = snapshot.get("version_id")
    item.parent_code = snapshot.get("parent_code")
    item.code = snapshot["code"]
    item.title_zh = snapshot.get("title_zh")
    item.title_en = snapshot.get("title_en") or snapshot["code"]
    item.level = snapshot.get("level", item.level)
    item.is_inspection = snapshot.get("is_inspection", item.is_inspection)
    item.before_sea_trial = snapshot.get("before_sea_trial", item.before_sea_trial)
    item.active = snapshot.get("active", item.active)
    item.sort_order = snapshot.get("sort_order", item.sort_order)
    item.updated_at = datetime.utcnow()


def audit_snapshot(value: str | None) -> dict | None:
    return json.loads(value) if value else None


def snapshot_ship(ship: Ship) -> dict:
    return {"id": ship.id, "project_id": ship.project_id, "hull_no": ship.hull_no, "name": ship.name}


def snapshot_progress(progress: ShipProgress) -> dict:
    return {
        "id": progress.id,
        "ship_id": progress.ship_id,
        "itp_item_id": progress.itp_item_id,
        "status": progress.status,
        "notes": progress.notes,
        "updated_by": progress.updated_by,
        "revision": progress.revision,
        "last_event_id": progress.last_event_id,
    }


def apply_progress_snapshots(progress: ShipProgress, item: ItpItem) -> None:
    progress.itp_item_id = item.id
    progress.item_uid = item.item_uid
    progress.version_id = item.version_id
    progress.code_snapshot = item.code
    progress.parent_code_snapshot = item.parent_code
    progress.title_zh_snapshot = item.title_zh
    progress.title_en_snapshot = item.title_en


def collect_descendant_items(db: Session, item: ItpItem) -> list[ItpItem]:
    all_items = db.query(ItpItem).filter(ItpItem.project_id == item.project_id).all()
    children_by_parent: dict[int | None, list[ItpItem]] = {}
    for candidate in all_items:
        children_by_parent.setdefault(candidate.parent_id, []).append(candidate)
    result = [item]
    stack = [item.id]
    while stack:
        parent_id = stack.pop()
        for child in children_by_parent.get(parent_id, []):
            result.append(child)
            stack.append(child.id)
    return result


def style_export_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="FFE2E8F0")
    alternate_fill = PatternFill("solid", fgColor="FFEAF4FF")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=1):
        if row_index % 2 == 1:
            for cell in row:
                cell.fill = alternate_fill


@app.get("/api/projects", response_model=list[ProjectOut])
def list_projects(db: Session = Depends(get_db)):
    return db.query(Project).order_by(Project.name).all()


@app.post("/api/projects", response_model=ProjectOut, dependencies=[Depends(require_admin)])
def create_project(payload: ProjectCreate, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    project = Project(name=payload.name.strip(), description=payload.description)
    db.add(project)
    db.flush()
    write_audit(
        db,
        entity_type="project",
        entity_id=project.id,
        action="create",
        summary=f"Created project {project.name}",
        actor=current_actor,
        after={"id": project.id, "name": project.name},
    )
    db.commit()
    db.refresh(project)
    return project


@app.delete("/api/projects/{project_id}", dependencies=[Depends(require_admin)])
def delete_project(project_id: int, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found.")
    before = {
        "id": project.id,
        "name": project.name,
        "ship_count": len(project.ships),
        "itp_item_count": len(project.items),
    }
    db.delete(project)
    write_audit(
        db,
        entity_type="project",
        entity_id=project_id,
        action="delete",
        summary=f"Deleted project {before['name']} with {before['ship_count']} ships and {before['itp_item_count']} ITP items",
        actor=current_actor,
        before=before,
        after=None,
    )
    db.commit()
    return {"ok": True}


@app.get("/api/projects/{project_id}/tree", response_model=list[ItpItemOut])
def get_project_tree(project_id: int, include_inactive: bool = False, db: Session = Depends(get_db)):
    query = db.query(ItpItem).filter(ItpItem.project_id == project_id)
    if not include_inactive:
        query = query.filter(ItpItem.active == True)
    items = query.order_by(ItpItem.sort_order, ItpItem.code).all()
    return build_tree(items)


@app.get("/api/projects/{project_id}/export", dependencies=[Depends(require_admin)])
def export_project_itp(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found.")
    items = sorted(
        db.query(ItpItem).filter(ItpItem.project_id == project_id, ItpItem.active == True).all(),
        key=item_code_key,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "检验项目导入数据对象"
    headers = ["No.", "Parent Code", "Current Code", "Chinese Description", "English Description", "Item UID", "Items Before Sea Trial"]
    sheet.append(headers)

    for index, item in enumerate(items, start=1):
        sheet.append([index, item.parent_code, item.code, item.title_zh, item.title_en, item.item_uid, "Y" if item.before_sea_trial else ""])

    style_export_sheet(sheet)
    widths = [10, 22, 22, 34, 48, 40, 24]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, column_index).column_letter].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"{project.name} ITP.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@app.post("/api/itp-items", response_model=ItpItemOut, dependencies=[Depends(require_admin)])
def create_itp_item(payload: ItpItemCreate, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    parent = None
    if payload.parent_code:
        parent = db.query(ItpItem).filter(ItpItem.project_id == payload.project_id, ItpItem.code == payload.parent_code, ItpItem.active == True).one_or_none()
        if parent is None:
            raise HTTPException(status_code=400, detail="Parent code was not found in this project.")
    item = ItpItem(
        project_id=payload.project_id,
        parent_id=parent.id if parent else None,
        parent_code=payload.parent_code,
        code=payload.code.strip(),
        title_zh=payload.title_zh.strip() if payload.title_zh else None,
        title_en=payload.title_en.strip(),
        is_inspection=payload.is_inspection,
        sort_order=payload.sort_order,
        level=(parent.level + 1) if parent else 1,
    )
    db.add(item)
    db.flush()
    items = db.query(ItpItem).filter(ItpItem.project_id == payload.project_id).all()
    resolve_levels_and_leaf_flags(items)
    write_audit(
        db,
        entity_type="itp_item",
        entity_id=item.id,
        action="create",
        summary=f"Created ITP item {item.code}",
        actor=current_actor,
        after=snapshot_item(item),
    )
    create_itp_version_snapshot(db, item.project, current_actor, "manual_create")
    db.commit()
    db.refresh(item)
    return ItpItemOut.model_validate(item)


@app.put("/api/itp-items/{item_id}", response_model=ItpItemOut, dependencies=[Depends(require_admin)])
def update_itp_item(item_id: int, payload: ItpItemUpdate, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    item = db.get(ItpItem, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="ITP item not found.")
    before = snapshot_item(item)
    fields = payload.model_fields_set
    old_code = item.code

    if "code" in fields:
        next_code = (payload.code or "").strip()
        if not next_code:
            raise HTTPException(status_code=400, detail="Current code cannot be empty.")
        duplicate = (
            db.query(ItpItem)
            .filter(ItpItem.project_id == item.project_id, ItpItem.code == next_code, ItpItem.id != item.id)
            .first()
        )
        if duplicate is not None:
            raise HTTPException(status_code=400, detail="Current code already exists in this project.")
        item.code = next_code
    if "title_zh" in fields:
        item.title_zh = (payload.title_zh or "").strip() or None
    if "title_en" in fields:
        next_title_en = (payload.title_en or "").strip()
        if not next_title_en:
            raise HTTPException(status_code=400, detail="English description cannot be empty.")
        item.title_en = next_title_en
    if "parent_code" in fields:
        next_parent_code = (payload.parent_code or "").strip() or None
        if next_parent_code is None:
            item.parent_code = None
            item.parent_id = None
        else:
            parent = (
                db.query(ItpItem)
                .filter(ItpItem.project_id == item.project_id, ItpItem.code == next_parent_code, ItpItem.active == True)
                .one_or_none()
            )
            if parent is None:
                raise HTTPException(status_code=400, detail="Parent code was not found as an active item in this project.")
            descendant_ids = {row.id for row in collect_descendant_items(db, item)}
            if parent.id in descendant_ids:
                raise HTTPException(status_code=400, detail="An ITP item cannot be moved under itself or one of its descendants.")
            item.parent_code = next_parent_code
            item.parent_id = parent.id
            item.parent_uid = parent.item_uid
    if old_code != item.code:
        for child in db.query(ItpItem).filter(ItpItem.project_id == item.project_id, ItpItem.id != item.id).all():
            if child.parent_id == item.id or child.parent_code == old_code:
                child.parent_code = item.code
    if payload.is_inspection is not None:
        item.is_inspection = payload.is_inspection
    if payload.sort_order is not None:
        item.sort_order = payload.sort_order
    item.updated_at = datetime.utcnow()
    db.flush()
    items = db.query(ItpItem).filter(ItpItem.project_id == item.project_id).all()
    resolve_levels_and_leaf_flags(items)
    active_over_depth = [row.code for row in items if row.active and row.level > 5]
    if active_over_depth:
        raise HTTPException(status_code=400, detail=f"Move would create levels deeper than level 5: {', '.join(active_over_depth[:5])}")
    write_audit(
        db,
        entity_type="itp_item",
        entity_id=item.id,
        action="update",
        summary=f"Updated ITP item {item.code}",
        actor=current_actor,
        before=before,
        after=snapshot_item(item),
    )
    create_itp_version_snapshot(db, item.project, current_actor, "manual_update")
    db.commit()
    db.refresh(item)
    return ItpItemOut.model_validate(item)


@app.put("/api/itp-items/{item_id}/before-sea-trial", response_model=ItpItemOut, dependencies=[Depends(require_admin)])
def toggle_before_sea_trial(item_id: int, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    item = db.get(ItpItem, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="ITP item not found.")
    target_value = not item.before_sea_trial
    affected_items = collect_descendant_items(db, item)
    before = {"root": snapshot_item(item), "affected": [snapshot_item(row) for row in affected_items]}
    for affected in affected_items:
        affected.before_sea_trial = target_value
    db.flush()
    write_audit(
        db,
        entity_type="itp_item",
        entity_id=item.id,
        action="mark",
        summary=f"{'Marked' if target_value else 'Unmarked'} {len(affected_items)} ITP item(s) under {item.code} as Items before sea trial",
        actor=current_actor,
        before=before,
        after={"root": snapshot_item(item), "affected": [snapshot_item(row) for row in affected_items]},
    )
    create_itp_version_snapshot(db, item.project, current_actor, "before_sea_trial_mark")
    db.commit()
    db.refresh(item)
    return ItpItemOut.model_validate(item)


@app.put("/api/itp-items/{item_id}/active", response_model=ItpItemOut, dependencies=[Depends(require_admin)])
def set_itp_item_active(
    item_id: int,
    active: bool = Query(...),
    db: Session = Depends(get_db),
    current_actor: str = Depends(actor),
):
    item = db.get(ItpItem, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="ITP item not found.")
    affected_items = collect_descendant_items(db, item)
    before = {"root": snapshot_item(item), "affected": [snapshot_item(row) for row in affected_items]}
    for affected in affected_items:
        affected.active = active
        affected.updated_at = datetime.utcnow()
    db.flush()
    write_audit(
        db,
        entity_type="itp_item",
        entity_id=item.id,
        action="activate" if active else "deactivate",
        summary=f"{'Activated' if active else 'Deactivated'} {len(affected_items)} ITP item(s) under {item.code}",
        actor=current_actor,
        before=before,
        after={"root": snapshot_item(item), "affected": [snapshot_item(row) for row in affected_items]},
    )
    create_itp_version_snapshot(db, item.project, current_actor, "manual_active_change")
    db.commit()
    db.refresh(item)
    return ItpItemOut.model_validate(item)


def permanently_delete_itp_item_records(db: Session, item: ItpItem):
    item_uid = item.item_uid
    progress_query = db.query(ShipProgress).filter(ShipProgress.itp_item_id == item.id)
    event_query = db.query(ShipProgressEvent).filter(ShipProgressEvent.itp_item_id == item.id)
    if item_uid:
        progress_query = db.query(ShipProgress).filter((ShipProgress.itp_item_id == item.id) | (ShipProgress.item_uid == item_uid))
        event_query = db.query(ShipProgressEvent).filter((ShipProgressEvent.itp_item_id == item.id) | (ShipProgressEvent.item_uid == item_uid))
    progress_count = progress_query.delete(synchronize_session=False)
    event_count = event_query.delete(synchronize_session=False)
    version_item_count = 0
    if item_uid:
        version_item_count = db.query(ItpVersionItem).filter(ItpVersionItem.item_uid == item_uid).delete(synchronize_session=False)
    db.delete(item)
    db.flush()
    return {
        "progress_count": progress_count,
        "event_count": event_count,
        "version_item_count": version_item_count,
    }


@app.delete("/api/itp-items/{item_id}", dependencies=[Depends(require_admin)])
def delete_itp_item(item_id: int, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    item = db.get(ItpItem, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="ITP item not found.")
    has_children = db.query(ItpItem).filter(ItpItem.parent_id == item.id).first() is not None
    if has_children:
        raise HTTPException(status_code=400, detail="This ITP item has child items. Delete the child items first.")
    project_id = item.project_id
    before = snapshot_item(item)
    deleted = permanently_delete_itp_item_records(db, item)
    remaining = db.query(ItpItem).filter(ItpItem.project_id == project_id).all()
    resolve_levels_and_leaf_flags(remaining)
    write_audit(
        db,
        entity_type="itp_item",
        entity_id=item_id,
        action="delete_permanent",
        summary=f"Permanently deleted ITP item {before['code']} with {deleted['progress_count']} progress record(s), {deleted['event_count']} event(s), and {deleted['version_item_count']} version snapshot(s)",
        actor=current_actor,
        before={
            "id": before["id"],
            "project_id": before["project_id"],
            "code": before["code"],
            "title_zh": before["title_zh"],
            "title_en": before["title_en"],
            "level": before["level"],
            "progress_count": deleted["progress_count"],
            "event_count": deleted["event_count"],
            "version_item_count": deleted["version_item_count"],
        },
        after=None,
    )
    project = db.get(Project, project_id)
    if project:
        create_itp_version_snapshot(db, project, current_actor, "manual_delete")
    db.commit()
    return {"ok": True, "action": "delete_permanent", "affected": 1, "progress_deleted": deleted["progress_count"], "events_deleted": deleted["event_count"]}


@app.delete("/api/projects/{project_id}/inactive-items", dependencies=[Depends(require_admin)])
def delete_inactive_itp_items(project_id: int, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found.")
    inactive_items = db.query(ItpItem).filter(ItpItem.project_id == project_id, ItpItem.active == False).all()
    if not inactive_items:
        return {"ok": True, "action": "delete_inactive_permanent", "affected": 0, "progress_deleted": 0, "events_deleted": 0}

    blocked = []
    for item in inactive_items:
        has_active_child = db.query(ItpItem).filter(ItpItem.parent_id == item.id, ItpItem.active == True).first() is not None
        if has_active_child:
            blocked.append(item.code)
    if blocked:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot clear inactive items because these inactive items still have active child items: {', '.join(blocked[:10])}",
        )

    pending = sorted(inactive_items, key=lambda row: (row.level, row.code), reverse=True)
    before_items = [snapshot_item(item) for item in pending]
    totals = {"progress_count": 0, "event_count": 0, "version_item_count": 0}
    affected = 0
    while pending:
        next_pending = []
        deleted_any = False
        for item in pending:
            has_children = db.query(ItpItem).filter(ItpItem.parent_id == item.id).first() is not None
            if has_children:
                next_pending.append(item)
                continue
            deleted = permanently_delete_itp_item_records(db, item)
            totals["progress_count"] += deleted["progress_count"]
            totals["event_count"] += deleted["event_count"]
            totals["version_item_count"] += deleted["version_item_count"]
            affected += 1
            deleted_any = True
        if next_pending and not deleted_any:
            blocked_codes = ", ".join(item.code for item in next_pending[:10])
            raise HTTPException(status_code=400, detail=f"Cannot clear inactive items with child items: {blocked_codes}")
        pending = next_pending

    remaining = db.query(ItpItem).filter(ItpItem.project_id == project_id).all()
    resolve_levels_and_leaf_flags(remaining)
    write_audit(
        db,
        entity_type="itp_item",
        entity_id=project_id,
        action="delete_inactive_permanent",
        summary=f"Permanently deleted {affected} inactive ITP item(s) in {project.name}",
        actor=current_actor,
        before={
            "project_id": project_id,
            "project_name": project.name,
            "items": before_items,
            "progress_count": totals["progress_count"],
            "event_count": totals["event_count"],
            "version_item_count": totals["version_item_count"],
        },
        after=None,
    )
    create_itp_version_snapshot(db, project, current_actor, "manual_delete_inactive")
    db.commit()
    return {
        "ok": True,
        "action": "delete_inactive_permanent",
        "affected": affected,
        "progress_deleted": totals["progress_count"],
        "events_deleted": totals["event_count"],
    }


@app.post("/api/import/preview", response_model=ImportPreview, dependencies=[Depends(require_admin)])
async def import_preview(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    try:
        result = preview_import(db, content)
        db.rollback()
        return result
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception:
        db.rollback()
        raise


@app.post("/api/import/apply", response_model=ProjectOut, dependencies=[Depends(require_admin)])
async def import_apply(
    mode: str = Query(default="partial", pattern="^(partial|global)$"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_actor: str = Depends(actor),
):
    content = await file.read()
    try:
        return import_itp(db, content, current_actor, mode)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/projects/{project_id}/ships", response_model=list[ShipOut])
def list_ships(project_id: int, db: Session = Depends(get_db)):
    return db.query(Ship).filter(Ship.project_id == project_id).order_by(Ship.hull_no).all()


@app.post("/api/ships", response_model=ShipOut, dependencies=[Depends(require_admin)])
def create_ship(payload: ShipCreate, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    ship = Ship(project_id=payload.project_id, hull_no=payload.hull_no.strip(), name=payload.name)
    db.add(ship)
    db.flush()
    write_audit(
        db,
        entity_type="ship",
        entity_id=ship.id,
        action="create",
        summary=f"Created ship {ship.hull_no}",
        actor=current_actor,
        after=snapshot_ship(ship),
    )
    db.commit()
    db.refresh(ship)
    return ship


@app.put("/api/ships/{ship_id}", response_model=ShipOut, dependencies=[Depends(require_admin)])
def update_ship(ship_id: int, payload: ShipUpdate, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    ship = db.get(Ship, ship_id)
    if ship is None:
        raise HTTPException(status_code=404, detail="Ship not found.")
    before = snapshot_ship(ship)
    if payload.hull_no is not None:
        hull_no = payload.hull_no.strip()
        if not hull_no:
            raise HTTPException(status_code=400, detail="Hull no. cannot be empty.")
        conflict = db.query(Ship).filter(Ship.project_id == ship.project_id, Ship.hull_no == hull_no, Ship.id != ship.id).one_or_none()
        if conflict is not None:
            raise HTTPException(status_code=400, detail="Another ship already uses this hull number in the project.")
        ship.hull_no = hull_no
    if payload.name is not None:
        ship.name = payload.name.strip() or None
    db.flush()
    after = snapshot_ship(ship)
    if before != after:
        write_audit(
            db,
            entity_type="ship",
            entity_id=ship.id,
            action="update",
            summary=f"Updated ship {ship.hull_no}",
            actor=current_actor,
            before=before,
            after=after,
        )
    db.commit()
    db.refresh(ship)
    return ship


@app.delete("/api/ships/{ship_id}", dependencies=[Depends(require_admin)])
def delete_ship(ship_id: int, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    ship = db.get(Ship, ship_id)
    if ship is None:
        raise HTTPException(status_code=404, detail="Ship not found.")
    before = snapshot_ship(ship)
    progress_count = db.query(ShipProgress).filter(ShipProgress.ship_id == ship_id).count()
    db.delete(ship)
    write_audit(
        db,
        entity_type="ship",
        entity_id=ship_id,
        action="delete",
        summary=f"Deleted ship {before['hull_no']} and {progress_count} progress records",
        actor=current_actor,
        before={**before, "progress_count": progress_count},
        after=None,
    )
    db.commit()
    return {"ok": True}


@app.get("/api/ships/{ship_id}/progress", response_model=list[ShipProgressOut])
def get_ship_progress(ship_id: int, db: Session = Depends(get_db)):
    ship = db.get(Ship, ship_id)
    if ship is None:
        raise HTTPException(status_code=404, detail="Ship not found.")
    items = sorted(
        db.query(ItpItem).filter(ItpItem.project_id == ship.project_id, ItpItem.is_inspection == True, ItpItem.active == True).all(),
        key=item_code_key,
    )
    progress_rows = db.query(ShipProgress).filter(ShipProgress.ship_id == ship_id).all()
    progress_by_uid = {row.item_uid: row for row in progress_rows if row.item_uid}
    progress_by_item = {row.itp_item_id: row for row in progress_rows}
    result = []
    for item in items:
        current = progress_by_uid.get(item.item_uid) or progress_by_item.get(item.id)
        result.append(
            ShipProgressOut(
                item_id=item.id,
                code=item.code,
                title_zh=item.title_zh,
                title_en=item.title_en,
                status=current.status if current else "not_started",
                notes=current.notes if current else None,
                updated_by=current.updated_by if current else None,
                updated_at=current.updated_at if current else None,
                completed_at=current.completed_at if current else None,
                revision=current.revision if current else 0,
            )
        )
    return result


@app.get("/api/ships/{ship_id}/records/export", dependencies=[Depends(require_admin)])
def export_ship_records(ship_id: int, db: Session = Depends(get_db)):
    ship = db.get(Ship, ship_id)
    if ship is None:
        raise HTTPException(status_code=404, detail="Ship not found.")
    project = db.get(Project, ship.project_id)
    items = sorted(
        db.query(ItpItem).filter(ItpItem.project_id == ship.project_id, ItpItem.is_inspection == True, ItpItem.active == True).all(),
        key=item_code_key,
    )
    progress_rows = db.query(ShipProgress).filter(ShipProgress.ship_id == ship_id).all()
    progress_by_uid = {row.item_uid: row for row in progress_rows if row.item_uid}
    progress_by_item = {row.itp_item_id: row for row in progress_rows}
    item_by_uid = {item.item_uid: item for item in items if item.item_uid}
    item_by_code = {item.code: item for item in items}

    workbook = Workbook()
    current_sheet = workbook.active
    current_sheet.title = "Current Status"
    current_headers = [
        "No.",
        "Project",
        "Hull No.",
        "Ship Name",
        "Item UID",
        "ITP Version ID",
        "Parent Code",
        "Current Code",
        "Chinese Description",
        "English Description",
        "Items Before Sea Trial",
        "Status",
        "Notes",
        "Updated By",
        "Updated At",
        "Completed At",
        "Revision",
    ]
    current_sheet.append(current_headers)

    for index, item in enumerate(items, start=1):
        current = progress_by_uid.get(item.item_uid) or progress_by_item.get(item.id)
        current_sheet.append(
            [
                index,
                project.name if project else "",
                ship.hull_no,
                ship.name,
                item.item_uid,
                item.version_id,
                item.parent_code,
                item.code,
                item.title_zh,
                item.title_en,
                "Y" if item.before_sea_trial else "",
                current.status if current else "not_started",
                current.notes if current else None,
                current.updated_by if current else None,
                current.updated_at if current else None,
                current.completed_at if current else None,
                current.revision if current else 0,
            ]
        )

    events_sheet = workbook.create_sheet("Status Events")
    event_headers = [
        "Event ID",
        "Created At",
        "Project",
        "Hull No.",
        "Ship Name",
        "Item UID",
        "ITP Version ID",
        "Parent Code Snapshot",
        "Code Snapshot",
        "Chinese Description Snapshot",
        "English Description Snapshot",
        "Items Before Sea Trial",
        "Status Before",
        "Status After",
        "Notes",
        "Updated By",
    ]
    events_sheet.append(event_headers)
    events = db.query(ShipProgressEvent).filter(ShipProgressEvent.ship_id == ship_id).order_by(ShipProgressEvent.created_at, ShipProgressEvent.id).all()
    for event in events:
        event_item = item_by_uid.get(event.item_uid) or item_by_code.get(event.code_snapshot)
        events_sheet.append(
            [
                event.id,
                event.created_at,
                project.name if project else "",
                ship.hull_no,
                ship.name,
                event.item_uid,
                event.version_id,
                event.parent_code_snapshot,
                event.code_snapshot,
                event.title_zh_snapshot,
                event.title_en_snapshot,
                "Y" if event_item and event_item.before_sea_trial else "",
                event.status_before,
                event.status_after,
                event.notes,
                event.updated_by,
            ]
        )

    for sheet in [current_sheet, events_sheet]:
        style_export_sheet(sheet)
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"{ship.hull_no} Inspection Records.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@app.post("/api/ships/{ship_id}/records/import", dependencies=[Depends(require_admin)])
async def import_ship_records(
    ship_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_actor: str = Depends(actor),
):
    ship = db.get(Ship, ship_id)
    if ship is None:
        raise HTTPException(status_code=404, detail="Ship not found.")

    workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    sheet = workbook["Current Status"] if "Current Status" in workbook.sheetnames else workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    index = {name: idx for idx, name in enumerate(headers)}
    required = ["Item UID", "Current Code", "Status"]
    missing = [name for name in required if name not in index]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    valid_statuses = {"not_started", "in_progress", "done", "not_applicable"}
    items = db.query(ItpItem).filter(ItpItem.project_id == ship.project_id, ItpItem.is_inspection == True, ItpItem.active == True).all()
    item_by_uid = {item.item_uid: item for item in items if item.item_uid}
    item_by_code = {item.code: item for item in items}
    imported = 0
    skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        item_uid = str(row[index["Item UID"]]).strip() if row[index["Item UID"]] is not None else ""
        code = str(row[index["Current Code"]]).strip() if row[index["Current Code"]] is not None else ""
        status = str(row[index["Status"]]).strip() if row[index["Status"]] is not None else ""
        if not item_uid and not code and not status:
            continue
        if status not in valid_statuses:
            skipped += 1
            continue
        item = item_by_uid.get(item_uid) or item_by_code.get(code)
        if item is None:
            skipped += 1
            continue

        notes = None
        if "Notes" in index and row[index["Notes"]] is not None:
            notes = str(row[index["Notes"]]).strip() or None

        progress = db.query(ShipProgress).filter(ShipProgress.ship_id == ship_id, ShipProgress.item_uid == item.item_uid).one_or_none()
        progress = progress or db.query(ShipProgress).filter(ShipProgress.ship_id == ship_id, ShipProgress.itp_item_id == item.id).one_or_none()
        before = snapshot_progress(progress) if progress else None
        if progress is None:
            progress = ShipProgress(ship_id=ship_id, itp_item_id=item.id)
            db.add(progress)
            current_revision = 0
            status_before = None
        else:
            current_revision = progress.revision
            status_before = progress.status

        apply_progress_snapshots(progress, item)
        progress.status = status
        progress.notes = notes
        progress.updated_by = current_actor
        progress.completed_at = datetime.utcnow() if status == "done" else None
        progress.revision = current_revision + 1
        db.flush()

        event = ShipProgressEvent(
            ship_id=ship.id,
            project_id=ship.project_id,
            itp_item_id=item.id,
            item_uid=item.item_uid,
            version_id=item.version_id,
            status_before=status_before,
            status_after=status,
            notes=notes,
            code_snapshot=item.code,
            parent_code_snapshot=item.parent_code,
            title_zh_snapshot=item.title_zh,
            title_en_snapshot=item.title_en,
            updated_by=current_actor,
        )
        db.add(event)
        db.flush()
        progress.last_event_id = event.id
        write_audit(
            db,
            entity_type="ship_progress",
            entity_id=progress.id,
            action="import",
            summary=f"Imported progress for ship {ship.hull_no} / item {item.code}",
            actor=current_actor,
            before=before,
            after=snapshot_progress(progress),
        )
        imported += 1

    db.commit()
    return {"ok": True, "imported": imported, "skipped": skipped}


@app.put("/api/ships/{ship_id}/progress/{item_id}", response_model=ShipProgressOut)
def update_ship_progress(
    ship_id: int,
    item_id: int,
    payload: ProgressUpdate,
    db: Session = Depends(get_db),
    current_actor: str = Depends(actor),
):
    if payload.status not in {"not_started", "in_progress", "done", "not_applicable"}:
        raise HTTPException(status_code=400, detail="Invalid status.")
    ship = db.get(Ship, ship_id)
    item = db.get(ItpItem, item_id)
    if ship is None or item is None or item.project_id != ship.project_id or not item.active:
        raise HTTPException(status_code=404, detail="Ship or ITP item not found.")
    progress = db.query(ShipProgress).filter(ShipProgress.ship_id == ship_id, ShipProgress.item_uid == item.item_uid).one_or_none()
    progress = progress or db.query(ShipProgress).filter(ShipProgress.ship_id == ship_id, ShipProgress.itp_item_id == item_id).one_or_none()
    before = None
    if progress is None:
        progress = ShipProgress(ship_id=ship_id, itp_item_id=item_id)
        db.add(progress)
    else:
        before = snapshot_progress(progress)
    current_revision = progress.revision if progress.id else 0
    if payload.expected_revision is not None and payload.expected_revision != current_revision:
        raise HTTPException(status_code=409, detail="This inspection item was updated by another user. Please refresh and try again.")
    status_before = progress.status if progress.id else None
    apply_progress_snapshots(progress, item)
    progress.status = payload.status
    progress.notes = payload.notes
    progress.updated_by = payload.updated_by
    progress.completed_at = datetime.utcnow() if payload.status == "done" else None
    progress.revision = current_revision + 1
    db.flush()
    event = ShipProgressEvent(
            ship_id=ship.id,
            project_id=ship.project_id,
            itp_item_id=item.id,
            item_uid=item.item_uid,
            version_id=item.version_id,
            status_before=status_before,
            status_after=progress.status,
            notes=progress.notes,
            code_snapshot=item.code,
            parent_code_snapshot=item.parent_code,
            title_zh_snapshot=item.title_zh,
            title_en_snapshot=item.title_en,
            updated_by=payload.updated_by,
    )
    db.add(event)
    db.flush()
    progress.last_event_id = event.id
    write_audit(
        db,
        entity_type="ship_progress",
        entity_id=progress.id,
        action="update",
        summary=f"Updated progress for ship {ship.hull_no} / item {item.code}",
        actor=current_actor,
        before=before,
        after=snapshot_progress(progress),
    )
    db.commit()
    db.refresh(progress)
    return ShipProgressOut(
        item_id=item.id,
        code=item.code,
        title_zh=item.title_zh,
        title_en=item.title_en,
        status=progress.status,
        notes=progress.notes,
        updated_by=progress.updated_by,
        updated_at=progress.updated_at,
        completed_at=progress.completed_at,
        revision=progress.revision,
    )


@app.get("/api/history", response_model=list[AuditLogOut])
def history(project_id: int | None = None, limit: int = 100, db: Session = Depends(get_db)):
    if not project_id:
        return db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(min(limit, 500)).all()

    project_item_ids = {row.id for row in db.query(ItpItem.id).filter(ItpItem.project_id == project_id).all()}
    ship_ids = {row.id for row in db.query(Ship.id).filter(Ship.project_id == project_id).all()}
    progress_ids = {
        row.id
        for row in db.query(ShipProgress.id)
        .filter((ShipProgress.ship_id.in_(ship_ids or {-1})) | (ShipProgress.itp_item_id.in_(project_item_ids or {-1})))
        .all()
    }

    def belongs_to_project(row: AuditLog) -> bool:
        before = audit_snapshot(row.before_json) or {}
        after = audit_snapshot(row.after_json) or {}
        snapshots = [before, after]
        if row.entity_type == "project":
            return row.entity_id == project_id or any(snapshot.get("id") == project_id for snapshot in snapshots)
        if row.entity_type == "itp_item":
            return row.entity_id in project_item_ids or any(snapshot.get("project_id") == project_id for snapshot in snapshots)
        if row.entity_type == "ship":
            return row.entity_id in ship_ids or any(snapshot.get("project_id") == project_id for snapshot in snapshots)
        if row.entity_type == "ship_progress":
            return (
                row.entity_id in progress_ids
                or any(snapshot.get("ship_id") in ship_ids for snapshot in snapshots)
                or any(snapshot.get("itp_item_id") in project_item_ids for snapshot in snapshots)
            )
        return False

    rows = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(2000).all()
    filtered = [row for row in rows if belongs_to_project(row)]
    return filtered[: min(limit, 500)]


@app.get("/api/overview")
def overview(db: Session = Depends(get_db)):
    projects = db.query(Project).order_by(Project.name).all()
    rows = []
    ship_rows = []
    for project in projects:
        item_ids = [row.id for row in db.query(ItpItem.id).filter(ItpItem.project_id == project.id, ItpItem.is_inspection == True, ItpItem.active == True).all()]
        ship_ids = [row.id for row in db.query(Ship.id).filter(Ship.project_id == project.id).all()]
        done_count = 0
        total_slots = len(item_ids) * len(ship_ids)
        if item_ids and ship_ids:
            done_count = (
                db.query(ShipProgress)
                .filter(
                    ShipProgress.ship_id.in_(ship_ids),
                    ShipProgress.itp_item_id.in_(item_ids),
                    ShipProgress.status == "done",
                )
                .count()
            )
        rows.append(
            {
                "project_id": project.id,
                "project_name": project.name,
                "ship_count": len(ship_ids),
                "itp_item_count": len(item_ids),
                "completion_done": done_count,
                "completion_total": total_slots,
            }
        )
        project_ships = db.query(Ship).filter(Ship.project_id == project.id).order_by(Ship.hull_no).all()
        item_uids = [
            row.item_uid
            for row in db.query(ItpItem.item_uid).filter(
                ItpItem.project_id == project.id,
                ItpItem.is_inspection == True,
                ItpItem.active == True,
            ).all()
        ]
        before_sea_trial_item_ids = [
            row.id
            for row in db.query(ItpItem.id).filter(
                ItpItem.project_id == project.id,
                ItpItem.is_inspection == True,
                ItpItem.active == True,
                ItpItem.before_sea_trial == True,
            ).all()
        ]
        before_sea_trial_item_uids = [
            row.item_uid
            for row in db.query(ItpItem.item_uid).filter(
                ItpItem.project_id == project.id,
                ItpItem.is_inspection == True,
                ItpItem.active == True,
                ItpItem.before_sea_trial == True,
            ).all()
        ]
        for ship in project_ships:
            ship_done = 0
            before_sea_trial_done = 0
            if item_ids:
                ship_done = (
                    db.query(ShipProgress)
                    .filter(
                        ShipProgress.ship_id == ship.id,
                        (ShipProgress.itp_item_id.in_(item_ids)) | (ShipProgress.item_uid.in_(item_uids or ["__none__"])),
                        ShipProgress.status == "done",
                    )
                    .count()
                )
            if before_sea_trial_item_ids:
                before_sea_trial_done = (
                    db.query(ShipProgress)
                    .filter(
                        ShipProgress.ship_id == ship.id,
                        (ShipProgress.itp_item_id.in_(before_sea_trial_item_ids)) | (ShipProgress.item_uid.in_(before_sea_trial_item_uids or ["__none__"])),
                        ShipProgress.status == "done",
                    )
                    .count()
                )
            ship_rows.append(
                {
                    "project_id": project.id,
                    "project_name": project.name,
                    "ship_id": ship.id,
                    "hull_no": ship.hull_no,
                    "ship_name": ship.name,
                    "completion_done": ship_done,
                    "completion_total": len(item_ids),
                    "completion_open": max(len(item_ids) - ship_done, 0),
                    "completion_percent": round((ship_done / len(item_ids)) * 100) if item_ids else 0,
                    "before_sea_trial_done": before_sea_trial_done,
                    "before_sea_trial_total": len(before_sea_trial_item_ids),
                    "before_sea_trial_open": max(len(before_sea_trial_item_ids) - before_sea_trial_done, 0),
                    "before_sea_trial_percent": round((before_sea_trial_done / len(before_sea_trial_item_ids)) * 100) if before_sea_trial_item_ids else 0,
                }
            )
    return {
        "project_count": len(projects),
        "ship_count": db.query(Ship).count(),
        "itp_item_count": db.query(ItpItem).filter(ItpItem.is_inspection == True, ItpItem.active == True).count(),
        "history_count": db.query(AuditLog).count(),
        "projects": rows,
        "ships": ship_rows,
    }


OPEN_ITEM_SCOPES = {
    "before_sea_trial": {
        "title": "Before Sea Trial Open Items",
        "summary_total_label": "Total Before Sea Trial",
        "empty_text": "No open items before sea trial.",
        "filename_scope": "Before Sea Trial Open Items",
    },
    "before_delivery": {
        "title": "Before Delivery Open Items",
        "summary_total_label": "Total Before Delivery",
        "empty_text": "No open items before delivery.",
        "filename_scope": "Before Delivery Open Items",
    },
}


def ship_open_items_data(ship_id: int, db: Session, scope: str) -> dict:
    ship = db.get(Ship, ship_id)
    if ship is None:
        raise HTTPException(status_code=404, detail="Ship not found.")
    scope_config = OPEN_ITEM_SCOPES.get(scope)
    if scope_config is None:
        raise HTTPException(status_code=400, detail="Invalid open item scope.")
    project = db.get(Project, ship.project_id)
    project_items = sorted(
        db.query(ItpItem).filter(ItpItem.project_id == ship.project_id, ItpItem.active == True).all(),
        key=item_code_key,
    )
    items_by_id = {item.id: item for item in project_items}
    inspection_items = [
        item
        for item in project_items
        if item.is_inspection and item.level == 5 and (scope != "before_sea_trial" or item.before_sea_trial)
    ]
    progress_rows = db.query(ShipProgress).filter(ShipProgress.ship_id == ship_id).all()
    progress_by_uid = {row.item_uid: row for row in progress_rows if row.item_uid}
    progress_by_item = {row.itp_item_id: row for row in progress_rows}

    def ancestor_for(item: ItpItem, level: int) -> ItpItem | None:
        current = item
        seen: set[int] = set()
        while current and current.id not in seen:
            if current.level == level:
                return current
            seen.add(current.id)
            current = items_by_id.get(current.parent_id) if current.parent_id else None
        return None

    grouped: dict[str, dict] = {}
    done = 0
    for item in inspection_items:
        progress = progress_by_uid.get(item.item_uid) or progress_by_item.get(item.id)
        if progress and progress.status == "done":
            done += 1
            continue
        level3 = ancestor_for(item, 3)
        level4 = ancestor_for(item, 4)
        level3_key = level3.code if level3 else "Ungrouped"
        group = grouped.setdefault(
            level3_key,
            {
                "code": level3.code if level3 else "Ungrouped",
                "title_en": level3.title_en if level3 else "Ungrouped",
                "title_zh": level3.title_zh if level3 else None,
                "groups": {},
            },
        )
        level4_key = level4.code if level4 else "Ungrouped"
        subgroup = group["groups"].setdefault(
            level4_key,
            {
                "code": level4.code if level4 else "Ungrouped",
                "title_en": level4.title_en if level4 else "Ungrouped",
                "title_zh": level4.title_zh if level4 else None,
                "items": [],
            },
        )
        subgroup["items"].append(
            {
                "id": item.id,
                "code": item.code,
                "title_en": item.title_en,
                "title_zh": item.title_zh,
                "status": progress.status if progress else "not_started",
            }
        )

    groups = []
    for group in grouped.values():
        subgroups = list(group["groups"].values())
        for subgroup in subgroups:
            subgroup["open_count"] = len(subgroup["items"])
        group["groups"] = subgroups
        group["open_count"] = sum(subgroup["open_count"] for subgroup in subgroups)
        groups.append(group)

    total = len(inspection_items)
    open_count = sum(group["open_count"] for group in groups)
    return {
        "project_id": ship.project_id,
        "project_name": project.name if project else "",
        "ship_id": ship.id,
        "hull_no": ship.hull_no,
        "ship_name": ship.name,
        "scope": scope,
        "title": scope_config["title"],
        "summary_total_label": scope_config["summary_total_label"],
        "empty_text": scope_config["empty_text"],
        "filename_scope": scope_config["filename_scope"],
        "done": done,
        "total": total,
        "open": open_count,
        "percent": round((done / total) * 100) if total else 0,
        "groups": groups,
    }


def before_sea_trial_open_items_data(ship_id: int, db: Session) -> dict:
    return ship_open_items_data(ship_id, db, "before_sea_trial")


def before_delivery_open_items_data(ship_id: int, db: Session) -> dict:
    return ship_open_items_data(ship_id, db, "before_delivery")


@app.get("/api/ships/{ship_id}/unfinished-before-sea-trial")
def unfinished_before_sea_trial(ship_id: int, db: Session = Depends(get_db)):
    return before_sea_trial_open_items_data(ship_id, db)


@app.get("/api/ships/{ship_id}/unfinished-before-delivery")
def unfinished_before_delivery(ship_id: int, db: Session = Depends(get_db)):
    return before_delivery_open_items_data(ship_id, db)


def pdf_paragraph(value: str | None, style: ParagraphStyle) -> Paragraph:
    text = str(value or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return Paragraph(text, style)


def add_pdf_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont(pdf_font_name(), 8)
    canvas.setFillColor(colors.HexColor("#64748b"))
    canvas.drawString(18 * mm, 12 * mm, "JN VLEC Project ITP Database - PG Newbuilding")
    canvas.drawRightString(A4[0] - 18 * mm, 12 * mm, f"Page {doc.page}")
    canvas.restoreState()


def build_open_items_pdf(data: dict) -> BytesIO:
    font_name = pdf_font_name()
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
        title=f"{data['hull_no']} {data['title']}",
    )

    base_styles = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "ITPTitle",
            parent=base_styles["Title"],
            fontName=font_name,
            fontSize=21,
            leading=25,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#0f172a"),
            spaceAfter=6,
        ),
        "hull": ParagraphStyle(
            "ITPHull",
            parent=base_styles["Title"],
            fontName=font_name,
            fontSize=18,
            leading=22,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#174ea6"),
        ),
        "ship_name": ParagraphStyle(
            "ITPShipName",
            parent=base_styles["Normal"],
            fontName=font_name,
            fontSize=14,
            leading=18,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#334155"),
        ),
        "label": ParagraphStyle(
            "ITPLabel",
            parent=base_styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#64748b"),
        ),
        "project": ParagraphStyle(
            "ITPProject",
            parent=base_styles["Normal"],
            fontName=font_name,
            fontSize=15,
            leading=19,
            textColor=colors.HexColor("#0f172a"),
        ),
        "subtitle": ParagraphStyle(
            "ITPSubtitle",
            parent=base_styles["Normal"],
            fontName=font_name,
            fontSize=9,
            leading=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#475569"),
            spaceAfter=12,
        ),
        "section": ParagraphStyle(
            "ITPSection",
            parent=base_styles["Heading2"],
            fontName=font_name,
            fontSize=12,
            leading=15,
            textColor=colors.HexColor("#0f766e"),
            spaceBefore=8,
            spaceAfter=5,
        ),
        "subsection": ParagraphStyle(
            "ITPSubsection",
            parent=base_styles["Heading3"],
            fontName=font_name,
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#334155"),
            spaceBefore=5,
            spaceAfter=3,
        ),
        "cell": ParagraphStyle(
            "ITPCell",
            parent=base_styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#1f2937"),
        ),
        "header": ParagraphStyle(
            "ITPHeaderCell",
            parent=base_styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#1e3a8a"),
        ),
    }

    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    logo_path = pg_logo_path()
    logo = ReportLabImage(logo_path, width=28 * mm, height=24 * mm) if logo_path else ""
    header = Table(
        [
            [
                logo,
                [
                    pdf_paragraph(data["title"], styles["title"]),
                    pdf_paragraph(f"Exported at {exported_at}", styles["subtitle"]),
                ],
            ]
        ],
        colWidths=[34 * mm, 142 * mm],
    )
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    ship_name = data.get("ship_name") or "Unnamed ship"
    info_card = Table(
        [
            [
                [pdf_paragraph("PROJECT", styles["label"]), pdf_paragraph(data["project_name"], styles["project"])],
                [pdf_paragraph(data["hull_no"], styles["hull"]), pdf_paragraph(ship_name, styles["ship_name"])],
            ]
        ],
        colWidths=[88 * mm, 88 * mm],
    )
    info_card.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eff6ff")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#bfdbfe")),
                ("LINEBEFORE", (1, 0), (1, 0), 0.5, colors.HexColor("#bfdbfe")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )

    story = [header, info_card, Spacer(1, 10)]

    summary = Table(
        [
            ["Open Items", data["summary_total_label"], "Completed", "Completion"],
            [str(data["open"]), str(data["total"]), str(data["done"]), f"{data['percent']}%"],
        ],
        colWidths=[38 * mm, 48 * mm, 38 * mm, 38 * mm],
        hAlign="CENTER",
    )
    summary.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#ecfeff")),
                ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#0f172a")),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bae6fd")),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend([summary, Spacer(1, 8)])

    if not data["groups"]:
        story.append(pdf_paragraph(data["empty_text"], styles["section"]))
    for group in data["groups"]:
        story.append(pdf_paragraph(f"{group['code']} - {group['title_en']} ({group['open_count']} open)", styles["section"]))
        for subgroup in group["groups"]:
            story.append(pdf_paragraph(f"{subgroup['code']} - {subgroup['title_en']} ({subgroup['open_count']} open)", styles["subsection"]))
            rows = [
                [
                    pdf_paragraph("Code", styles["header"]),
                    pdf_paragraph("English Description", styles["header"]),
                    pdf_paragraph("Chinese Description", styles["header"]),
                    pdf_paragraph("Status", styles["header"]),
                ]
            ]
            for item in subgroup["items"]:
                rows.append(
                    [
                        pdf_paragraph(item["code"], styles["cell"]),
                        pdf_paragraph(item["title_en"], styles["cell"]),
                        pdf_paragraph(item.get("title_zh"), styles["cell"]),
                        pdf_paragraph(item["status"].replace("_", " ").title(), styles["cell"]),
                    ]
                )
            table = Table(rows, colWidths=[24 * mm, 70 * mm, 58 * mm, 24 * mm], repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        ("LEFTPADDING", (0, 0), (-1, -1), 5),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eff6ff")]),
                    ]
                )
            )
            story.extend([table, Spacer(1, 6)])

    doc.build(story, onFirstPage=add_pdf_footer, onLaterPages=add_pdf_footer)
    output.seek(0)
    return output


def open_items_pdf_response(data: dict) -> StreamingResponse:
    output = build_open_items_pdf(data)
    filename = f"{data['hull_no']} {data['filename_scope']}.pdf"
    encoded = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@app.get("/api/ships/{ship_id}/unfinished-before-sea-trial/export.pdf")
def export_unfinished_before_sea_trial_pdf(ship_id: int, db: Session = Depends(get_db)):
    return open_items_pdf_response(before_sea_trial_open_items_data(ship_id, db))


@app.get("/api/ships/{ship_id}/unfinished-before-delivery/export.pdf")
def export_unfinished_before_delivery_pdf(ship_id: int, db: Session = Depends(get_db)):
    return open_items_pdf_response(before_delivery_open_items_data(ship_id, db))


@app.post("/api/history/{audit_id}/rollback", dependencies=[Depends(require_admin)])
def rollback_history(audit_id: int, db: Session = Depends(get_db), current_actor: str = Depends(actor)):
    audit = db.get(AuditLog, audit_id)
    if audit is None:
        raise HTTPException(status_code=404, detail="History record not found.")
    if audit.action == "rollback":
        raise HTTPException(status_code=400, detail="Rollback records cannot be rolled back.")
    if audit.action == "delete_permanent":
        raise HTTPException(status_code=400, detail="Permanent delete records cannot be rolled back.")

    before = audit_snapshot(audit.before_json)
    after = audit_snapshot(audit.after_json)
    itp_project_for_version: Project | None = None

    if audit.entity_type == "itp_item":
        item = db.get(ItpItem, audit.entity_id)
        if audit.action == "create":
            if item is None:
                raise HTTPException(status_code=404, detail="ITP item already removed.")
            has_children = db.query(ItpItem).filter(ItpItem.parent_id == item.id).first() is not None
            has_progress = db.query(ShipProgress).filter((ShipProgress.itp_item_id == item.id) | (ShipProgress.item_uid == item.item_uid)).first() is not None
            if has_children or has_progress:
                raise HTTPException(status_code=400, detail="This ITP item has children or progress records. Roll those back first.")
            project_id = item.project_id
            current = snapshot_item(item)
            db.delete(item)
            db.flush()
            remaining = db.query(ItpItem).filter(ItpItem.project_id == project_id).all()
            resolve_levels_and_leaf_flags(remaining)
            itp_project_for_version = db.get(Project, project_id)
            write_audit(
                db,
                entity_type="itp_item",
                entity_id=audit.entity_id,
                action="rollback",
                summary=f"Rolled back creation of ITP item {current['code']}",
                actor=current_actor,
                before=current,
                after=None,
            )
        elif audit.action == "delete" and before:
            if item is not None:
                raise HTTPException(status_code=400, detail="ITP item already exists.")
            conflict = db.query(ItpItem).filter(ItpItem.project_id == before["project_id"], ItpItem.code == before["code"]).one_or_none()
            if conflict is not None:
                raise HTTPException(status_code=400, detail="Another ITP item already uses this code.")
            item = ItpItem(id=before["id"])
            apply_item_snapshot(item, before)
            db.add(item)
            db.flush()
            items = db.query(ItpItem).filter(ItpItem.project_id == item.project_id).all()
            resolve_levels_and_leaf_flags(items)
            itp_project_for_version = db.get(Project, item.project_id)
            write_audit(
                db,
                entity_type="itp_item",
                entity_id=item.id,
                action="rollback",
                summary=f"Restored deleted ITP item {item.code} from history #{audit.id}",
                actor=current_actor,
                before=None,
                after=snapshot_item(item),
            )
        elif audit.action in {"mark", "activate", "deactivate"} and before and before.get("affected"):
            restored = []
            current = []
            for snapshot in before["affected"]:
                affected = db.get(ItpItem, snapshot["id"])
                if affected is None:
                    continue
                current.append(snapshot_item(affected))
                affected.before_sea_trial = snapshot.get("before_sea_trial", False)
                affected.active = snapshot.get("active", affected.active)
                affected.updated_at = datetime.utcnow()
                restored.append(snapshot_item(affected))
                if itp_project_for_version is None:
                    itp_project_for_version = affected.project
            if not restored:
                raise HTTPException(status_code=400, detail="No affected ITP items still exist for this marker rollback.")
            write_audit(
                db,
                entity_type="itp_item",
                entity_id=audit.entity_id,
                action="rollback",
                summary=f"Rolled back batch ITP item change from history #{audit.id}",
                actor=current_actor,
                before={"affected": current},
                after={"affected": restored},
            )
        elif before:
            if before.get("affected"):
                raise HTTPException(status_code=400, detail="This ITP history record is a batch change and cannot be restored as a single item.")
            if item is None:
                item = ItpItem(id=before["id"])
                db.add(item)
                db.flush()
            current = snapshot_item(item)
            apply_item_snapshot(item, before)
            db.flush()
            items = db.query(ItpItem).filter(ItpItem.project_id == item.project_id).all()
            resolve_levels_and_leaf_flags(items)
            itp_project_for_version = db.get(Project, item.project_id)
            write_audit(
                db,
                entity_type="itp_item",
                entity_id=item.id,
                action="rollback",
                summary=f"Rolled back ITP item {item.code} to history #{audit.id}",
                actor=current_actor,
                before=current,
                after=snapshot_item(item),
            )
        else:
            raise HTTPException(status_code=400, detail="This ITP history record has no rollback snapshot.")

    elif audit.entity_type == "ship_progress":
        progress = db.get(ShipProgress, audit.entity_id)
        if audit.action == "create":
            if progress is None:
                raise HTTPException(status_code=404, detail="Progress record already removed.")
            current = {"status": progress.status, "notes": progress.notes, "updated_by": progress.updated_by}
            db.delete(progress)
            write_audit(
                db,
                entity_type="ship_progress",
                entity_id=audit.entity_id,
                action="rollback",
                summary=f"Rolled back creation of progress record #{audit.entity_id}",
                actor=current_actor,
                before=current,
                after=None,
            )
        elif before:
            if progress is None:
                raise HTTPException(status_code=404, detail="Progress record not found.")
            current = snapshot_progress(progress)
            progress.status = before.get("status", "not_started")
            progress.notes = before.get("notes")
            progress.updated_by = before.get("updated_by", current_actor)
            progress.completed_at = datetime.utcnow() if progress.status == "done" else None
            write_audit(
                db,
                entity_type="ship_progress",
                entity_id=progress.id,
                action="rollback",
                summary=f"Rolled back progress record #{progress.id} to history #{audit.id}",
                actor=current_actor,
                before=current,
                after=snapshot_progress(progress),
            )
        elif progress is not None:
            current = snapshot_progress(progress)
            db.delete(progress)
            write_audit(
                db,
                entity_type="ship_progress",
                entity_id=audit.entity_id,
                action="rollback",
                summary=f"Rolled back first progress update record #{audit.entity_id}",
                actor=current_actor,
                before=current,
                after=None,
            )
        else:
            raise HTTPException(status_code=400, detail="This progress history record has no rollback snapshot.")

    elif audit.entity_type == "ship" and audit.action == "create":
        ship = db.get(Ship, audit.entity_id)
        if ship is None:
            raise HTTPException(status_code=404, detail="Ship already removed.")
        current = snapshot_ship(ship)
        db.delete(ship)
        write_audit(
            db,
            entity_type="ship",
            entity_id=audit.entity_id,
            action="rollback",
            summary=f"Rolled back creation of ship {current['hull_no']}",
            actor=current_actor,
            before=current,
            after=None,
        )

    elif audit.entity_type == "ship" and audit.action == "update" and before:
        ship = db.get(Ship, audit.entity_id)
        if ship is None:
            raise HTTPException(status_code=404, detail="Ship not found.")
        conflict = db.query(Ship).filter(Ship.project_id == before["project_id"], Ship.hull_no == before["hull_no"], Ship.id != ship.id).one_or_none()
        if conflict is not None:
            raise HTTPException(status_code=400, detail="Another ship already uses the previous hull number.")
        current = snapshot_ship(ship)
        ship.project_id = before["project_id"]
        ship.hull_no = before["hull_no"]
        ship.name = before.get("name")
        db.flush()
        write_audit(
            db,
            entity_type="ship",
            entity_id=ship.id,
            action="rollback",
            summary=f"Rolled back ship {ship.hull_no} to history #{audit.id}",
            actor=current_actor,
            before=current,
            after=snapshot_ship(ship),
        )

    elif audit.entity_type == "ship" and audit.action == "delete" and before:
        if db.get(Ship, audit.entity_id) is not None:
            raise HTTPException(status_code=400, detail="Ship already exists.")
        project = db.get(Project, before["project_id"])
        if project is None:
            raise HTTPException(status_code=400, detail="The ship's project no longer exists.")
        conflict = db.query(Ship).filter(Ship.project_id == before["project_id"], Ship.hull_no == before["hull_no"]).one_or_none()
        if conflict is not None:
            raise HTTPException(status_code=400, detail="Another ship already uses this hull number in the project.")
        ship = Ship(id=before["id"], project_id=before["project_id"], hull_no=before["hull_no"], name=before.get("name"))
        db.add(ship)
        db.flush()
        write_audit(
            db,
            entity_type="ship",
            entity_id=ship.id,
            action="rollback",
            summary=f"Restored deleted ship {ship.hull_no} from history #{audit.id}",
            actor=current_actor,
            before=None,
            after=snapshot_ship(ship),
        )

    elif audit.entity_type == "project" and audit.action == "create":
        project = db.get(Project, audit.entity_id)
        if project is None:
            raise HTTPException(status_code=404, detail="Project already removed.")
        if project.items or project.ships:
            raise HTTPException(status_code=400, detail="Project has ITP items or ships. Roll those back first.")
        current = {"id": project.id, "name": project.name}
        db.delete(project)
        write_audit(
            db,
            entity_type="project",
            entity_id=audit.entity_id,
            action="rollback",
            summary=f"Rolled back creation of project {current['name']}",
            actor=current_actor,
            before=current,
            after=None,
        )
    else:
        raise HTTPException(status_code=400, detail="This history record is not rollback-capable yet.")

    if itp_project_for_version is not None:
        create_itp_version_snapshot(db, itp_project_for_version, current_actor, "rollback")

    db.commit()
    return {"ok": True}
